from __future__ import annotations

import csv
import html
import json
import os
import re
import shutil
import string
import subprocess
import tempfile
import urllib.parse
import urllib.request
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

from evolva.tools.base import ToolResult

TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".tsv", ".json", ".jsonl", ".xml", ".py", ".log", ".html", ".htm"}
SPREADSHEET_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm", ".parquet"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".pptx"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff"}
AUDIO_EXTENSIONS = {".mp3", ".m4a", ".wav", ".flac", ".ogg"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".mkv", ".webm", ".avi"}
ARCHIVE_EXTENSIONS = {".zip"}
TASKSET_EXTERNAL_BINARIES = ["tesseract", "ffmpeg", "ffprobe", "yt-dlp", "whisper", "pdftotext"]
TASKSET_BROWSER_SEARCH_BINARIES = ["node", "npx", "uv", "uvx", "playwright", "google-chrome", "chromium"]
TASKSET_SEARCH_ENV_KEYS = ["TAVILY_API_KEY", "BRAVE_API_KEY", "SERPAPI_API_KEY", "BING_SEARCH_API_KEY", "EXA_API_KEY"]
TASKSET_OPTIONAL_MODULES = ["pypdf", "PyPDF2", "pandas", "pyarrow", "openpyxl", "PIL", "pytesseract", "moviepy"]


def _module_available(name: str) -> bool:
    try:
        __import__(name)
        return True
    except Exception:
        return False


def _bounded(value: str, max_chars: int) -> tuple[str, bool]:
    limit = max(0, int(max_chars))
    if len(value) <= limit:
        return value, False
    return value[:limit], True


def _run_external(argv: list[str], timeout: int = 60, max_chars: int = 20000) -> dict[str, Any]:
    try:
        proc = subprocess.run(argv, text=True, capture_output=True, timeout=max(1, int(timeout)), shell=False)
    except FileNotFoundError:
        return {"ok": False, "returncode": None, "stdout": "", "stderr": f"Executable not found: {argv[0]}", "argv": argv, "truncated": False}
    except subprocess.TimeoutExpired as exc:
        stdout = (exc.stdout or "") if isinstance(exc.stdout, str) else (exc.stdout or b"").decode("utf-8", errors="replace")
        stderr = (exc.stderr or "") if isinstance(exc.stderr, str) else (exc.stderr or b"").decode("utf-8", errors="replace")
        output, truncated = _bounded((stdout + "\n" + stderr).strip(), max_chars)
        return {"ok": False, "returncode": None, "stdout": output, "stderr": f"Timed out after {timeout}s", "argv": argv, "timeout": timeout, "truncated": truncated}
    stdout, stdout_truncated = _bounded(proc.stdout or "", max_chars)
    stderr, stderr_truncated = _bounded(proc.stderr or "", max_chars)
    return {"ok": proc.returncode == 0, "returncode": proc.returncode, "stdout": stdout, "stderr": stderr, "argv": argv, "truncated": stdout_truncated or stderr_truncated}


def _missing_dependency(tool: str, install_hint: str, data: dict[str, Any] | None = None) -> ToolResult:
    payload = dict(data or {})
    payload.update({"status": "missing_dependency", "tool": tool, "install_hint": install_hint})
    return ToolResult(False, f"Missing optional dependency `{tool}`. {install_hint}", payload)


def normalize_answer(answer: str) -> str:
    """Normalize task-set final answers for lightweight exact matching."""

    text = str(answer or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    text = text.strip(string.whitespace + string.punctuation)
    return text


def _mcp_web_config_summary(mcp_config_file: str | Path | None = None) -> dict[str, Any]:
    names: list[str] = []
    matching: list[str] = []
    if mcp_config_file:
        path = Path(mcp_config_file).expanduser()
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                raw_servers = data.get("servers", data if isinstance(data, dict) else {})
                if isinstance(raw_servers, dict):
                    names = sorted(str(name) for name in raw_servers)
                    needles = ("browser", "playwright", "search", "brave", "tavily", "fetch", "web")
                    for name, item in raw_servers.items():
                        blob = f"{name} {json.dumps(item, ensure_ascii=False)}".lower()
                        if any(needle in blob for needle in needles):
                            matching.append(str(name))
            except Exception:
                pass
    return {"config_file": str(mcp_config_file) if mcp_config_file else "", "servers": names, "browser_search_servers": sorted(matching)}


def browser_search_health(mcp_config_file: str | Path | None = None) -> dict[str, Any]:
    binaries = {name: {"available": shutil.which(name) is not None, "path": shutil.which(name)} for name in TASKSET_BROWSER_SEARCH_BINARIES}
    env_keys = {key: {"configured": bool(os.getenv(key))} for key in TASKSET_SEARCH_ENV_KEYS}
    mcp_config = _mcp_web_config_summary(mcp_config_file)
    capabilities = {
        "static_web_fetch": True,
        "duckduckgo_html_search": True,
        "api_search": any(item["configured"] for item in env_keys.values()),
        "browser_mcp_configured": bool(mcp_config["browser_search_servers"]),
        "browser_runtime": binaries["npx"]["available"] or binaries["node"]["available"] or binaries["playwright"]["available"],
    }
    return {"binaries": binaries, "env_keys": env_keys, "mcp": mcp_config, "capabilities": capabilities}


def taskset_tool_health(mcp_config_file: str | Path | None = None) -> ToolResult:
    """Report optional local tools that materially improve task-set coverage."""

    binaries = {name: {"available": shutil.which(name) is not None, "path": shutil.which(name)} for name in TASKSET_EXTERNAL_BINARIES}
    modules = {name: {"available": _module_available(name)} for name in TASKSET_OPTIONAL_MODULES}
    capabilities = {
        "static_web_fetch": True,
        "text_csv_docx_pptx_zip": True,
        "xlsx": modules["openpyxl"]["available"] or True,
        "pdf_text": modules["pypdf"]["available"] or modules["PyPDF2"]["available"] or binaries["pdftotext"]["available"],
        "ocr_image": modules["PIL"]["available"] and modules["pytesseract"]["available"] or binaries["tesseract"]["available"],
        "audio_transcription": binaries["whisper"]["available"],
        "video_probe": binaries["ffprobe"]["available"],
        "video_frames": binaries["ffmpeg"]["available"],
        "youtube_metadata": binaries["yt-dlp"]["available"],
        "parquet_table": modules["pandas"]["available"] and (modules["pyarrow"]["available"] or True),
    }
    web_health = browser_search_health(mcp_config_file)
    capabilities.update({
        "web_search_api": web_health["capabilities"]["api_search"],
        "browser_mcp": web_health["capabilities"]["browser_mcp_configured"],
        "browser_runtime": web_health["capabilities"]["browser_runtime"],
    })
    missing_for_broad_coverage = [
        name
        for name in ["pdf_text", "ocr_image", "audio_transcription", "video_probe", "video_frames", "youtube_metadata", "parquet_table"]
        if not capabilities.get(name)
    ]
    lines = ["Task-set optional tool health", "Binaries:"]
    for name, info in binaries.items():
        lines.append(f"- {name}: {'ok' if info['available'] else 'missing'}" + (f" ({info['path']})" if info.get("path") else ""))
    lines.append("Python modules:")
    for name, info in modules.items():
        lines.append(f"- {name}: {'ok' if info['available'] else 'missing'}")
    lines.append("Capabilities:")
    for name, available in capabilities.items():
        lines.append(f"- {name}: {'ok' if available else 'missing'}")
    lines.append("Browser/search:")
    for name, info in web_health["binaries"].items():
        lines.append(f"- {name}: {'ok' if info['available'] else 'missing'}" + (f" ({info['path']})" if info.get("path") else ""))
    configured_keys = [key for key, info in web_health["env_keys"].items() if info["configured"]]
    lines.append(f"- search_api_keys: {', '.join(configured_keys) if configured_keys else 'none'}")
    mcp_servers = web_health["mcp"].get("browser_search_servers", [])
    lines.append(f"- browser_search_mcp_servers: {', '.join(mcp_servers) if mcp_servers else 'none'}")
    if missing_for_broad_coverage:
        lines.append("Recommended optional installs for stronger task-set coverage: brew install tesseract ffmpeg yt-dlp poppler; pip install pypdf openpyxl pandas pyarrow pillow pytesseract openai-whisper")
    return ToolResult(True, "\n".join(lines), {"binaries": binaries, "modules": modules, "browser_search": web_health, "capabilities": capabilities, "missing_for_broad_coverage": missing_for_broad_coverage})


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self._skip_depth += 1
        if tag.lower() in {"p", "br", "div", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self._skip_depth:
            self._skip_depth -= 1
        if tag.lower() in {"p", "div", "li", "tr", "h1", "h2", "h3"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._skip_depth:
            self.parts.append(data)

    def text(self) -> str:
        raw = html.unescape(" ".join(self.parts))
        lines = [re.sub(r"\s+", " ", line).strip() for line in raw.splitlines()]
        return "\n".join(line for line in lines if line)


def html_to_text(markup: str) -> str:
    parser = _HTMLTextExtractor()
    parser.feed(markup)
    return parser.text()



def _json_post(url: str, payload: dict[str, Any], headers: dict[str, str], timeout: int) -> dict[str, Any]:
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json", **headers}, method="POST")
    with urllib.request.urlopen(req, timeout=max(1, int(timeout))) as resp:
        return json.loads(resp.read().decode("utf-8", errors="replace"))


def _json_get(url: str, params: dict[str, Any], headers: dict[str, str], timeout: int) -> dict[str, Any]:
    req = urllib.request.Request(url + "?" + urllib.parse.urlencode(params), headers=headers)
    with urllib.request.urlopen(req, timeout=max(1, int(timeout))) as resp:
        return json.loads(resp.read().decode("utf-8", errors="replace"))


def _clean_html(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<.*?>", " ", html.unescape(str(value or "")))).strip()


def _normalize_search_rows(rows: list[dict[str, Any]], max_results: int, source: str) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for row in rows:
        url = str(row.get("url") or row.get("link") or row.get("href") or "").strip()
        title = _clean_html(str(row.get("title") or row.get("name") or url))
        snippet = _clean_html(str(row.get("snippet") or row.get("description") or row.get("content") or ""))
        if not url:
            continue
        normalized.append({"title": title, "url": url, "snippet": snippet, "source": source})
        if len(normalized) >= max_results:
            break
    return normalized


def _search_duckduckgo(query: str, max_results: int, timeout: int) -> list[dict[str, str]]:
    url = "https://duckduckgo.com/html/?" + urllib.parse.urlencode({"q": query})
    req = urllib.request.Request(url, headers={"User-Agent": "evolva-agent/0.1"})
    with urllib.request.urlopen(req, timeout=max(1, int(timeout))) as resp:
        page = resp.read().decode("utf-8", errors="replace")
    matches = re.findall(r'<a rel="nofollow" class="result__a" href="([^"]+)">(.*?)</a>', page)
    snippets = re.findall(r'<a class="result__snippet"[^>]*>(.*?)</a>|<div class="result__snippet"[^>]*>(.*?)</div>', page, flags=re.S)
    rows = []
    for idx, (href, title) in enumerate(matches):
        href = html.unescape(href)
        if "uddg=" in href:
            parsed = urllib.parse.urlparse(href)
            qs = urllib.parse.parse_qs(parsed.query)
            href = qs.get("uddg", [href])[0]
        snippet = ""
        if idx < len(snippets):
            snippet = next((part for part in snippets[idx] if part), "")
        rows.append({"title": _clean_html(title), "url": href, "snippet": _clean_html(snippet), "source": "duckduckgo"})
        if len(rows) >= max_results:
            break
    return rows


def web_search_pro(query: str, provider: str = "auto", max_results: int = 5, timeout: int = 15) -> ToolResult:
    """Search the web via configured API providers with DuckDuckGo HTML fallback."""

    query = str(query or "").strip()
    if not query:
        return ToolResult(False, "query is required", {"provider": provider})
    provider = str(provider or "auto").strip().lower()
    max_results = max(1, min(int(max_results), 20))
    tried: list[str] = []
    errors: dict[str, str] = {}

    def run_provider(name: str) -> list[dict[str, str]]:
        tried.append(name)
        if name == "tavily":
            key = os.getenv("TAVILY_API_KEY")
            if not key:
                raise RuntimeError("TAVILY_API_KEY is not configured")
            data = _json_post("https://api.tavily.com/search", {"api_key": key, "query": query, "max_results": max_results}, {}, timeout)
            return _normalize_search_rows(list(data.get("results", [])), max_results, "tavily")
        if name == "brave":
            key = os.getenv("BRAVE_API_KEY")
            if not key:
                raise RuntimeError("BRAVE_API_KEY is not configured")
            data = _json_get("https://api.search.brave.com/res/v1/web/search", {"q": query, "count": max_results}, {"Accept": "application/json", "X-Subscription-Token": key}, timeout)
            return _normalize_search_rows(list(data.get("web", {}).get("results", [])), max_results, "brave")
        if name == "serpapi":
            key = os.getenv("SERPAPI_API_KEY")
            if not key:
                raise RuntimeError("SERPAPI_API_KEY is not configured")
            data = _json_get("https://serpapi.com/search.json", {"q": query, "api_key": key, "num": max_results}, {}, timeout)
            return _normalize_search_rows(list(data.get("organic_results", [])), max_results, "serpapi")
        if name in {"duckduckgo", "ddg"}:
            return _search_duckduckgo(query, max_results, timeout)
        raise RuntimeError(f"Unknown provider: {name}")

    providers = [provider]
    if provider == "auto":
        providers = []
        if os.getenv("TAVILY_API_KEY"):
            providers.append("tavily")
        if os.getenv("BRAVE_API_KEY"):
            providers.append("brave")
        if os.getenv("SERPAPI_API_KEY"):
            providers.append("serpapi")
        providers.append("duckduckgo")
    for name in providers:
        try:
            rows = run_provider(name)
            if rows:
                return ToolResult(True, json.dumps(rows, ensure_ascii=False, indent=2), {"query": query, "provider": name, "tried": tried, "results": rows})
            errors[name] = "no results"
        except Exception as exc:
            errors[name] = str(exc)
            if provider != "auto":
                break
    return ToolResult(False, "Search failed: " + "; ".join(f"{k}: {v}" for k, v in errors.items()), {"query": query, "provider": provider, "tried": tried, "errors": errors})


def web_fetch(url: str, max_chars: int = 20000, timeout: int = 20) -> ToolResult:
    if not re.match(r"^https?://", str(url or ""), flags=re.I):
        return ToolResult(False, "web_fetch only supports http(s) URLs", {"url": url})
    req = urllib.request.Request(url, headers={"User-Agent": "evolva-agent/0.1"})
    try:
        with urllib.request.urlopen(req, timeout=max(1, int(timeout))) as resp:
            raw = resp.read(max(int(max_chars) * 4, 4096))
            content_type = resp.headers.get("content-type", "")
            final_url = resp.geturl()
            status = getattr(resp, "status", None)
    except Exception as exc:
        return ToolResult(False, f"Fetch failed: {exc}", {"url": url})
    text = raw.decode("utf-8", errors="replace")
    if "html" in content_type.lower() or "<html" in text[:500].lower():
        text = html_to_text(text)
    text = text[: int(max_chars)]
    return ToolResult(True, text, {"url": url, "final_url": final_url, "status": status, "content_type": content_type, "chars": len(text), "truncated": len(raw) >= int(max_chars) * 4})


def read_text_file(path: Path, max_chars: int = 20000) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return {"ok": True, "kind": "text", "text": text[:max_chars], "truncated": len(text) > max_chars}


def _xml_text(xml_text: str) -> str:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return re.sub(r"<[^>]+>", " ", xml_text)
    parts = [item.strip() for item in root.itertext() if item and item.strip()]
    return "\n".join(parts)


def _zip_member_text(zf: zipfile.ZipFile, name: str) -> str:
    try:
        return zf.read(name).decode("utf-8", errors="replace")
    except KeyError:
        return ""


def extract_docx(path: Path, max_chars: int = 20000) -> dict[str, Any]:
    with zipfile.ZipFile(path) as zf:
        text = _xml_text(_zip_member_text(zf, "word/document.xml"))
    return {"ok": True, "kind": "docx", "text": text[:max_chars], "truncated": len(text) > max_chars}


def extract_pptx(path: Path, max_chars: int = 20000) -> dict[str, Any]:
    parts: list[str] = []
    with zipfile.ZipFile(path) as zf:
        names = sorted(name for name in zf.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        for name in names:
            text = _xml_text(_zip_member_text(zf, name))
            if text:
                parts.append(f"[{Path(name).stem}]\n{text}")
    text = "\n\n".join(parts)
    return {"ok": True, "kind": "pptx", "text": text[:max_chars], "slides": len(parts), "truncated": len(text) > max_chars}


def extract_xlsx_fallback(path: Path, max_chars: int = 20000, max_rows: int = 20) -> dict[str, Any]:
    with zipfile.ZipFile(path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            try:
                root = ET.fromstring(_zip_member_text(zf, "xl/sharedStrings.xml"))
                for si in root.iter():
                    if si.tag.endswith("si"):
                        value = "".join(t.text or "" for t in si.iter() if t.tag.endswith("t"))
                        shared.append(value)
            except ET.ParseError:
                pass
        sheet_names = sorted(name for name in zf.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        lines: list[str] = []
        for sheet_name in sheet_names[:5]:
            lines.append(f"[{Path(sheet_name).stem}]")
            try:
                root = ET.fromstring(_zip_member_text(zf, sheet_name))
            except ET.ParseError:
                continue
            rows_seen = 0
            for row in root.iter():
                if not row.tag.endswith("row"):
                    continue
                values: list[str] = []
                for c in row:
                    if not c.tag.endswith("c"):
                        continue
                    cell_type = c.attrib.get("t")
                    v = next((child.text for child in c if child.tag.endswith("v")), "") or ""
                    if cell_type == "s" and v.isdigit() and int(v) < len(shared):
                        v = shared[int(v)]
                    values.append(v)
                if values:
                    lines.append("\t".join(values))
                    rows_seen += 1
                if rows_seen >= max_rows:
                    break
        text = "\n".join(lines)
    return {"ok": True, "kind": "xlsx", "text": text[:max_chars], "truncated": len(text) > max_chars, "engine": "zip-xml-fallback"}


def extract_xlsx(path: Path, max_chars: int = 20000, max_rows: int = 20) -> dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return extract_xlsx_fallback(path, max_chars=max_chars, max_rows=max_rows)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in wb.sheetnames[:10]:
        ws = wb[sheet_name]
        lines.append(f"[{sheet_name}]")
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                lines.append("\t".join(values).rstrip())
    text = "\n".join(lines)
    return {"ok": True, "kind": "xlsx", "text": text[:max_chars], "truncated": len(text) > max_chars, "engine": "openpyxl"}


def extract_pdf(path: Path, max_chars: int = 20000) -> dict[str, Any]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        try:
            from PyPDF2 import PdfReader  # type: ignore
        except Exception:
            return {"ok": False, "kind": "pdf", "text": "", "error": "PDF text extraction requires optional pypdf/PyPDF2 or an external Docling/MarkItDown/OCR tool."}
    try:
        reader = PdfReader(str(path))
        parts = []
        for page in reader.pages[:30]:
            parts.append(page.extract_text() or "")
        text = "\n".join(parts)
        return {"ok": True, "kind": "pdf", "text": text[:max_chars], "pages_previewed": min(len(reader.pages), 30), "truncated": len(text) > max_chars}
    except Exception as exc:
        return {"ok": False, "kind": "pdf", "text": "", "error": f"PDF extraction failed: {exc}"}


def pdf_extract_external(path: str | Path, max_chars: int = 20000, timeout: int = 60) -> ToolResult:
    """Extract PDF text with Python libraries first, then optional pdftotext CLI."""

    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    if p.suffix.lower() != ".pdf":
        return ToolResult(False, f"Not a PDF file: {p}", {"path": str(p), "extension": p.suffix.lower()})
    data = extract_pdf(p, max_chars=max_chars)
    if data.get("ok") and data.get("text"):
        data.update({"path": str(p), "engine": data.get("engine", "python-pdf")})
        return ToolResult(True, str(data.get("text", "")), data)
    if shutil.which("pdftotext"):
        run = _run_external(["pdftotext", "-layout", "-enc", "UTF-8", str(p), "-"], timeout=timeout, max_chars=max_chars)
        text = run.get("stdout", "")
        if run.get("ok") and text:
            return ToolResult(True, text, {"path": str(p), "kind": "pdf", "engine": "pdftotext", "text": text, "truncated": run.get("truncated", False), "process": run})
        return ToolResult(False, run.get("stderr") or "pdftotext failed", {"path": str(p), "kind": "pdf", "engine": "pdftotext", "process": run})
    return _missing_dependency("pypdf/PyPDF2 or pdftotext", "Install with `pip install pypdf` or `brew install poppler` for pdftotext.", {"path": str(p), "previous_error": data.get("error")})


def extract_zip(path: Path, max_chars: int = 20000) -> dict[str, Any]:
    lines: list[str] = []
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
        lines.append("Archive members:")
        lines.extend(f"- {name}" for name in names[:100])
        for name in names:
            suffix = Path(name).suffix.lower()
            if suffix in TEXT_EXTENSIONS and not name.endswith("/"):
                preview = zf.read(name).decode("utf-8", errors="replace")[:4000]
                lines.append(f"\n--- {name} ---\n{preview}")
            if len("\n".join(lines)) >= max_chars:
                break
    text = "\n".join(lines)
    return {"ok": True, "kind": "zip", "text": text[:max_chars], "truncated": len(text) > max_chars}


def image_metadata(path: Path) -> dict[str, Any]:
    data: dict[str, Any] = {"ok": True, "kind": "image", "text": f"Image file: {path.name} ({path.stat().st_size} bytes). OCR/image reasoning requires a vision model or OCR tool."}
    try:
        from PIL import Image  # type: ignore

        with Image.open(path) as image:
            data.update({"width": image.width, "height": image.height, "mode": image.mode, "format": image.format})
            data["text"] = f"Image file: {path.name}; {image.format} {image.width}x{image.height} {image.mode}. OCR/image reasoning requires a vision model or OCR tool."
    except Exception:
        pass
    return data


def ocr_image(path: str | Path, language: str = "eng", max_chars: int = 20000, timeout: int = 60) -> ToolResult:
    """OCR an image with optional pytesseract/Pillow or the tesseract CLI."""

    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    if p.suffix.lower() not in IMAGE_EXTENSIONS:
        return ToolResult(False, f"Not a supported image file: {p}", {"path": str(p), "extension": p.suffix.lower()})
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore

        with Image.open(p) as image:
            raw = pytesseract.image_to_string(image, lang=language or "eng")
        text, truncated = _bounded(raw, max_chars)
        return ToolResult(True, text, {"path": str(p), "kind": "ocr", "engine": "pytesseract", "language": language, "text": text, "truncated": truncated})
    except Exception as exc:
        py_error = str(exc)
    if shutil.which("tesseract"):
        run = _run_external(["tesseract", str(p), "stdout", "-l", language or "eng"], timeout=timeout, max_chars=max_chars)
        text = run.get("stdout", "")
        if run.get("ok"):
            return ToolResult(True, text, {"path": str(p), "kind": "ocr", "engine": "tesseract-cli", "language": language, "text": text, "truncated": run.get("truncated", False), "process": run})
        return ToolResult(False, run.get("stderr") or "tesseract failed", {"path": str(p), "kind": "ocr", "engine": "tesseract-cli", "process": run, "pytesseract_error": py_error})
    return _missing_dependency("tesseract", "Install `tesseract` (for example `brew install tesseract`) and optionally `pip install pillow pytesseract`.", {"path": str(p), "pytesseract_error": py_error})


def media_metadata(path: Path, kind: str) -> dict[str, Any]:
    need = "audio transcription" if kind == "audio" else "video frame/transcript extraction"
    return {"ok": True, "kind": kind, "text": f"{kind.title()} file: {path.name} ({path.stat().st_size} bytes). task solving may require {need} via Whisper/ffmpeg/yt-dlp or an external MCP tool."}


def audio_transcribe(path: str | Path, model: str = "base", language: str = "", max_chars: int = 20000, timeout: int = 600) -> ToolResult:
    """Transcribe audio/video through an installed Whisper CLI without downloading anything itself."""

    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    if p.suffix.lower() not in AUDIO_EXTENSIONS | VIDEO_EXTENSIONS:
        return ToolResult(False, f"Not a supported audio/video file: {p}", {"path": str(p), "extension": p.suffix.lower()})
    whisper = shutil.which("whisper")
    if not whisper:
        return _missing_dependency("whisper", "Install a local Whisper CLI, for example `pip install openai-whisper` (requires ffmpeg for many formats).", {"path": str(p)})
    with tempfile.TemporaryDirectory(prefix="evolva-taskset-whisper-") as tmp:
        argv = [whisper, str(p), "--model", model or "base", "--output_format", "txt", "--output_dir", tmp]
        if language:
            argv.extend(["--language", language])
        run = _run_external(argv, timeout=timeout, max_chars=8000)
        txt_files = sorted(Path(tmp).glob("*.txt"))
        text = "\n".join(item.read_text(encoding="utf-8", errors="replace") for item in txt_files)
    text, truncated = _bounded(text or run.get("stdout", ""), max_chars)
    ok = bool(text.strip()) and bool(run.get("ok"))
    data = {"path": str(p), "kind": "transcript", "engine": "whisper-cli", "model": model, "language": language, "text": text, "truncated": truncated or run.get("truncated", False), "process": run}
    return ToolResult(ok, text or run.get("stderr") or "whisper did not produce a transcript", data)


def video_probe(path: str | Path, timeout: int = 30) -> ToolResult:
    """Inspect video/audio stream metadata through ffprobe when available."""

    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    ffprobe = shutil.which("ffprobe")
    if not ffprobe:
        fallback = media_metadata(p, "video" if p.suffix.lower() in VIDEO_EXTENSIONS else "audio")
        return _missing_dependency("ffprobe", "Install with `brew install ffmpeg` to enable stream probing.", {"path": str(p), "fallback": fallback})
    run = _run_external([ffprobe, "-v", "error", "-print_format", "json", "-show_format", "-show_streams", str(p)], timeout=timeout, max_chars=40000)
    if not run.get("ok"):
        return ToolResult(False, run.get("stderr") or "ffprobe failed", {"path": str(p), "process": run})
    try:
        data = json.loads(run.get("stdout") or "{}")
    except json.JSONDecodeError:
        data = {"raw": run.get("stdout", "")}
    streams = data.get("streams") if isinstance(data, dict) else []
    fmt = data.get("format") if isinstance(data, dict) else {}
    duration = fmt.get("duration") if isinstance(fmt, dict) else None
    summary = f"{p.name}: duration={duration or 'unknown'}s streams={len(streams or [])}"
    return ToolResult(True, summary, {"path": str(p), "kind": "media_probe", "engine": "ffprobe", "probe": data, "process": run})


def video_extract_frames(path: str | Path, output_dir: str | Path, every_seconds: float = 10.0, max_frames: int = 12, timeout: int = 120) -> ToolResult:
    """Extract a bounded set of frames through ffmpeg into a caller-approved output directory."""

    p = Path(path).expanduser().resolve()
    out = Path(output_dir).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    if p.suffix.lower() not in VIDEO_EXTENSIONS:
        return ToolResult(False, f"Not a supported video file: {p}", {"path": str(p), "extension": p.suffix.lower()})
    ffmpeg = shutil.which("ffmpeg")
    if not ffmpeg:
        return _missing_dependency("ffmpeg", "Install with `brew install ffmpeg` to enable frame extraction.", {"path": str(p), "output_dir": str(out)})
    out.mkdir(parents=True, exist_ok=True)
    interval = max(0.1, float(every_seconds))
    frame_limit = max(1, min(int(max_frames), 100))
    pattern = out / "frame_%04d.jpg"
    argv = [ffmpeg, "-hide_banner", "-loglevel", "error", "-i", str(p), "-vf", f"fps=1/{interval}", "-frames:v", str(frame_limit), str(pattern)]
    run = _run_external(argv, timeout=timeout, max_chars=12000)
    frames = sorted(str(item) for item in out.glob("frame_*.jpg"))[:frame_limit]
    ok = bool(run.get("ok")) and bool(frames)
    output = "\n".join(frames) if frames else run.get("stderr") or "ffmpeg produced no frames"
    return ToolResult(ok, output, {"path": str(p), "kind": "video_frames", "engine": "ffmpeg", "output_dir": str(out), "frames": frames, "every_seconds": interval, "max_frames": frame_limit, "process": run})


def yt_dlp_info(url: str, max_chars: int = 30000, timeout: int = 120) -> ToolResult:
    """Fetch video/page metadata through yt-dlp when installed."""

    if not re.match(r"^https?://", str(url or ""), flags=re.I):
        return ToolResult(False, "yt_dlp_info only supports http(s) URLs", {"url": url})
    ytdlp = shutil.which("yt-dlp")
    if not ytdlp:
        return _missing_dependency("yt-dlp", "Install with `brew install yt-dlp` or `pipx install yt-dlp`.", {"url": url})
    run = _run_external([ytdlp, "--dump-json", "--skip-download", "--no-warnings", str(url)], timeout=timeout, max_chars=max_chars)
    if not run.get("ok"):
        return ToolResult(False, run.get("stderr") or "yt-dlp failed", {"url": url, "process": run})
    raw = run.get("stdout", "")
    try:
        data = json.loads(raw.splitlines()[0] if raw.splitlines() else "{}")
    except json.JSONDecodeError:
        data = {"raw": raw}
    summary = {key: data.get(key) for key in ["id", "title", "duration", "uploader", "webpage_url", "description"] if isinstance(data, dict) and key in data}
    text = json.dumps(summary or data, ensure_ascii=False, indent=2)[:max_chars]
    return ToolResult(True, text, {"url": url, "kind": "yt_dlp_info", "engine": "yt-dlp", "info": data, "process": run})


def parquet_preview(path: Path, max_chars: int = 20000, max_rows: int = 20) -> dict[str, Any]:
    try:
        import pandas as pd  # type: ignore
    except Exception:
        return {"ok": False, "kind": "parquet", "text": "", "error": "Parquet preview requires optional pandas/pyarrow/duckdb."}
    try:
        df = pd.read_parquet(path)
        text = df.head(max_rows).to_string(index=False)
        return {"ok": True, "kind": "parquet", "text": text[:max_chars], "rows": int(len(df)), "columns": list(map(str, df.columns)), "truncated": len(text) > max_chars}
    except Exception as exc:
        return {"ok": False, "kind": "parquet", "text": "", "error": f"Parquet preview failed: {exc}"}


def file_to_text(path: str | Path, max_chars: int = 20000, max_rows: int = 20) -> ToolResult:
    p = Path(path).expanduser().resolve()
    if not p.exists() or not p.is_file():
        return ToolResult(False, f"File not found: {p}", {"path": str(p)})
    suffix = p.suffix.lower()
    try:
        if suffix in TEXT_EXTENSIONS:
            data = read_text_file(p, max_chars=max_chars)
        elif suffix == ".docx":
            data = extract_docx(p, max_chars=max_chars)
        elif suffix == ".pptx":
            data = extract_pptx(p, max_chars=max_chars)
        elif suffix in {".xlsx", ".xlsm"}:
            data = extract_xlsx(p, max_chars=max_chars, max_rows=max_rows)
        elif suffix == ".pdf":
            data = extract_pdf(p, max_chars=max_chars)
        elif suffix == ".zip":
            data = extract_zip(p, max_chars=max_chars)
        elif suffix in IMAGE_EXTENSIONS:
            data = image_metadata(p)
        elif suffix in AUDIO_EXTENSIONS:
            data = media_metadata(p, "audio")
        elif suffix in VIDEO_EXTENSIONS:
            data = media_metadata(p, "video")
        elif suffix == ".parquet":
            data = parquet_preview(p, max_chars=max_chars, max_rows=max_rows)
        else:
            data = {"ok": False, "kind": suffix.lstrip(".") or "unknown", "text": "", "error": f"Unsupported file type: {suffix}"}
    except Exception as exc:
        data = {"ok": False, "kind": suffix.lstrip("."), "text": "", "error": str(exc)}
    data.update({"path": str(p), "extension": suffix, "size_bytes": p.stat().st_size})
    output = data.get("text") or data.get("error") or ""
    return ToolResult(bool(data.get("ok")), str(output), data)


def spreadsheet_describe(path: str | Path, max_rows: int = 20, max_chars: int = 20000) -> ToolResult:
    p = Path(path).expanduser().resolve()
    suffix = p.suffix.lower()
    if suffix not in SPREADSHEET_EXTENSIONS:
        return ToolResult(False, f"Not a supported spreadsheet/table file: {p}", {"path": str(p), "extension": suffix})
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        rows: list[list[str]] = []
        with p.open(newline="", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for idx, row in enumerate(reader):
                rows.append(row)
                if idx + 1 >= int(max_rows):
                    break
        text = "\n".join("\t".join(row) for row in rows)
        return ToolResult(True, text[:max_chars], {"path": str(p), "kind": suffix.lstrip("."), "rows_previewed": len(rows), "rows": rows})
    return file_to_text(p, max_chars=max_chars, max_rows=max_rows)


def classify_taskset_task(question: str, file_name: str = "", annotator_metadata: str = "") -> dict[str, Any]:
    text = f"{question or ''}\n{file_name or ''}\n{annotator_metadata or ''}".lower()
    suffix = Path(file_name).suffix.lower() if file_name else ""
    categories: list[str] = []
    if re.search(r"https?://|website|webpage|internet|google|search|wikipedia|youtube|online|site\b|url", text):
        categories.append("web")
    if re.search(r"calculate|how many|sum|average|ratio|percent|number|count|difference|total|compute", text):
        categories.append("calc")
    if suffix in SPREADSHEET_EXTENSIONS or re.search(r"spreadsheet|excel|csv|table|dataframe|parquet", text):
        categories.append("table")
    if suffix in DOCUMENT_EXTENSIONS or re.search(r"pdf|document|presentation|slides|docx|pptx", text):
        categories.append("document")
    if suffix in IMAGE_EXTENSIONS or re.search(r"image|picture|photo|screenshot|ocr|visual", text):
        categories.append("image")
    if suffix in AUDIO_EXTENSIONS or re.search(r"audio|listen|mp3|transcript", text):
        categories.append("audio")
    if suffix in VIDEO_EXTENSIONS or re.search(r"video|movie|youtube", text):
        categories.append("video")
    if suffix in {".py", ".json", ".xml"} or re.search(r"python|code|script|json|xml", text):
        categories.append("code")
    categories = list(dict.fromkeys(categories)) or ["text"]
    blockers = []
    if "web" in categories:
        blockers.append("dynamic browsing/search may need browser/search MCP for live pages")
    if "image" in categories:
        blockers.append("OCR/vision questions need a vision model or OCR tool")
    if "audio" in categories:
        blockers.append("audio questions need transcription tool")
    if "video" in categories:
        blockers.append("video questions need frame/transcript extraction")
    native = {"text", "calc", "table", "document", "code"}
    status = "native_or_likely" if set(categories) <= native else "partial_tooling"
    if {"audio", "video"} & set(categories):
        status = "needs_external_media_tool"
    return {"categories": categories, "status": status, "blockers": blockers, "file_extension": suffix}


def load_taskset_metadata(metadata_csv: str | Path) -> list[dict[str, str]]:
    path = Path(metadata_csv).expanduser().resolve()
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        return [dict(row) for row in csv.DictReader(f)]


def resolve_attachment(task_id: str, file_name: str, attachments_dir: str | Path, file_path: str = "") -> dict[str, Any]:
    if not file_name:
        return {"task_id": task_id, "file_name": file_name, "path": None, "exists": False, "reason": "no file_name"}
    base = Path(attachments_dir).expanduser().resolve()
    candidates = [base / file_name, base / str(task_id) / file_name]
    if file_path:
        rel = Path(file_path)
        if rel.is_absolute():
            candidates.append(rel)
        else:
            candidates.extend([base / rel, base.parent / rel, base.parent.parent / rel])
    matches = [candidate for candidate in candidates if candidate.exists()]
    if not matches:
        found = list(base.rglob(file_name))[:5] if base.exists() else []
        matches.extend(found)
    path = matches[0] if matches else candidates[0]
    return {"task_id": task_id, "file_name": file_name, "path": str(path), "exists": path.exists(), "size_bytes": path.stat().st_size if path.exists() else None}


def taskset_context(metadata_csv: str | Path, attachments_dir: str | Path, task_id: str = "", limit: int = 5, max_chars: int = 12000) -> ToolResult:
    rows = load_taskset_metadata(metadata_csv)
    if task_id:
        selected = [row for row in rows if row.get("task_id") == task_id]
    else:
        selected = rows[: int(limit)]
    contexts = []
    for row in selected:
        attachment = resolve_attachment(row.get("task_id", ""), row.get("file_name", ""), attachments_dir, row.get("file_path", ""))
        preview = None
        if attachment.get("exists") and attachment.get("path"):
            preview_result = file_to_text(attachment["path"], max_chars=max_chars)
            preview = {"ok": preview_result.ok, "output": preview_result.output[:max_chars], "data": preview_result.data}
        contexts.append({"task_id": row.get("task_id"), "level": row.get("Level"), "question": row.get("Question"), "file_name": row.get("file_name"), "attachment": attachment, "classification": classify_taskset_task(row.get("Question", ""), row.get("file_name", ""), row.get("Annotator Metadata", "")), "preview": preview})
    output = json.dumps(contexts, ensure_ascii=False, indent=2)
    return ToolResult(True, output, {"count": len(contexts), "tasks": contexts})
